from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

arquivo = Workbook()
planilha = arquivo.active
planilha.title = "Produtos"

# negrito, cor letra
config_fonte_cabecalho = Font(bold=True, color="FFFFFF")
# cor fundo, preechimento
config_fundo_cabecalho = PatternFill(fgColor="4f4f4f", fill_type="solid")

planilha.merge_cells("A1:D1")
planilha['A1'] = "Relatório de Vendas"
# alinhamento do texto (centralizado)
planilha['A1'].alignment = Alignment(horizontal="center")
planilha['A1'].font = Font(bold=True, size=14, color="FFFFFF")
# 
planilha['A1'].fill = PatternFill(fgColor="1F497D", fill_type="solid")

# implementando as estilizações percorrendo linha
planilha.append(['Produto', 'Preço', 'Quantidade', 'Data'])
for celula in planilha[2]:
    celula.font = config_fonte_cabecalho
    celula.fill = config_fundo_cabecalho

agora = datetime.now()

planilha.append(['Camiseta', 59.99, 10, agora])
planilha.append(['Calça Jeans', 120, 5, agora])
planilha.append(['Tênis', 250, 2, agora])

# percorrendo 
for celula in planilha['B']:
    if celula.row <= 2:
        continue
    celula.alignment = Alignment(horizontal="right")
    celula.number_format = "R$ #,##0.00"

for celula in planilha['D']:
    celula.number_format = "DD/MM/YY"

# pra escrvaer em pt
fina = Side(style="thin")
# estilo da borda
borda = Border(right=fina,left=fina, top=fina, bottom=fina)
for linha in planilha.iter_rows():
    for celula in linha:
        celula.border = borda

arquivo.save("Produtos.xlsx")