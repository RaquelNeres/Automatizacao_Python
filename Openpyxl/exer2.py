from openpyxl import load_workbook

# 1
arquivo = load_workbook('Openpyxl/alunos.xlsx')
planilha_alunos = arquivo['Alunos']

print(planilha_alunos['B2'].value)
print(planilha_alunos['D5'].value)
print(planilha_alunos['E10'].value)

# 2
for coluna in planilha_alunos['D']:
    if coluna.row == 1:
        continue

    if float(coluna.value) >= 8.0:
        print(coluna.value)

# 3
# iterar nas linhas, so considera os valores das celulas, linha que ele começa, linha que ele termina
for linha in planilha_alunos.iter_rows(values_only=True, min_row=2):
    print('-'*50)
    # pega de uma tupla (a linha retorna uma tupla)
    aluno, curso, idade, nota, matricula = linha
    print(f"""ALUNO: {aluno}
CURSO: {curso}
IDADE: {idade}
NOTA FINAL: {nota}
MATRÍCULA: {matricula.strftime("%d/%m/%Y")}""")