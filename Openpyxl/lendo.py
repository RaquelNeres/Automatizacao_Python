from openpyxl import load_workbook

# abrindo planilha que ja existe
arquivo = load_workbook('Openpyxl/planilha_funcionarios.xlsx')
planilha_funcionarios = arquivo['Funcionários']

# salario_anthony = planilha_funcionarios['D6'].value
# print(salario_anthony) # .value

# todos os valores da linha
linha_13 = planilha_funcionarios[13]
for celula in linha_13:
    print(celula.value)

linhas = planilha_funcionarios[7:12]
for linha in linhas:
    print("="*30)
    for celula in linha:
        print(celula.value)

coluna_salario = planilha_funcionarios['D']
print(coluna_salario)

# iterar nas linhas, so considera os valores das celulas, linha que ele começa, linha que ele termina
for linha in planilha_funcionarios.iter_rows(values_only=True, min_row=2, max_row=20):
    print('-'*50)
    # pega de uma tupla (a linha retorna uma tupla)
    nome, departamento, idade, salario, data_admissao = linha
    print(f"""NOME: {nome}
DEPARTAMENTO: {departamento}
IDADE: {idade}
SALÁRIO: {salario}
DATA DE ADMISSÃO: {data_admissao.strftime("%d/%m/%Y")}""")