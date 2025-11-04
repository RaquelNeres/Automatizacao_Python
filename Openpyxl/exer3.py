from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

#criando
arquivo = Workbook()
diario = arquivo.active
diario.title = "Diario"

# estrutura
diario.merge_cells("A1:D1")
diario["A1"] = "Diário de Leituras – Agosto 2025"
diario.append(['Livro', 'Autor', 'Data de Início', 'Processo(%)'])

# livros
diario.append(['O Peregrino', 'John Bunyan', '12/04/2024', 1.0])
diario.append(['O Homem mais Inteligente da História', 'Augusto Cury', '05/06/2024', 0.5])
diario.append(['A Cabana', 'William P. Young', '22/07/2024', 1.0])
diario.append(['O Monge e o Executivo', 'James C. Hunter', '10/08/2024', 1.0])
diario.append(['Python para Iniciantes', 'Michael Dawson', '02/01/2025', 0.5])
diario.append(['Clean Code', 'Robert C. Martin', '15/02/2025', 0.3])
diario.append(['Ame de Verdade', 'Max Lucado', '01/03/2025', 0.5])
diario.append(['Inteligência Artificial: Estruturas e Estratégias', 'George Luger', '18/04/2025', 0.2])
diario.append(['A Arte da Guerra', 'Sun Tzu', '20/05/2025', 1.0])
diario.append(['Cristianismo Puro e Simples', 'C.S. Lewis', '01/06/2025', 0.5])
diario.append(['O Pequeno Príncipe', 'Antoine de Saint-Exupéry', '10/06/2025', 1.0])
diario.append(['O Código Da Vinci', 'Dan Brown', '15/06/2025', 0.4])
diario.append(['A Revolução dos Bichos', 'George Orwell', '01/07/2025', 1.0])
diario.append(['O Poder do Hábito', 'Charles Duhigg', '05/07/2025', 0.6])
diario.append(['A Bíblia do Programador', 'Eric S. Raymond', '12/07/2025', 0.25])

# titulo
diario['A1'].alignment = Alignment(horizontal='center')
diario['A1'].fill = PatternFill(fgColor="1F497D", fill_type="solid")
diario['A1'].font = Font(bold=True, size=14, color='FFFFFF')

fina = Side(style="thin")
# estilo da borda
borda = Border(right=fina,left=fina, top=fina, bottom=fina)

# cabeçalho
for bloco in diario[2]:
    bloco.font = Font(bold=True, color='FFFFFF')
    bloco.fill = PatternFill(fgColor='4f4f4f',fill_type='solid')
    bloco.border = borda

# linhas dos livros
cont = 0
for linha in diario.iter_rows(min_row=3):
    for celula in linha:
        if (cont%2) != 0:
            celula.fill = PatternFill(fgColor='4f4f4f', fill_type='solid')
        celula.border = borda
    cont+=1

for coluna in diario['A'][2:]:
    coluna.alignment = Alignment(horizontal='left')

for coluna in diario['B'][2:]:
    coluna.alignment = Alignment(horizontal='left')

for coluna in diario['C'][2:]:
    coluna.alignment = Alignment(horizontal='center')
    celula.number_format = 'DD/MM/YYYY'

for coluna in diario['D'][2:]:
    coluna.alignment = Alignment(horizontal='right')
    coluna.number_format = ".0%"


arquivo.save('Openpyxl/planilhas_criadas/diario_leituras.xlsx')
