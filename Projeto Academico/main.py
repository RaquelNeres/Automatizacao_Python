from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

arquivo = load_workbook('Projeto Academico/alunos.xlsx')
alunos = arquivo['Alunos']

re = Workbook()
reprovados = re.active
reprovados.title = "Reprovados"

ap = Workbook()
aprovados = ap.active
aprovados.title = "Aprovados"

apro = 0
repro = 0
total = 0
cont = 0
maior = 1
nome = ""

aprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])
reprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])

for linha in alunos.iter_rows(values_only=True, min_row=2, max_row=20):
    Nome, Curso, Idade, Nota, Matrícula = linha
    total += Nota
    cont += 1
    if Nota > maior:
        maior = Nota
        nome = Nome

    if Nota >= 7:
        aprovados.append([Nome, Curso, Idade, Nota, Matrícula])
        apro+=1
    else:
        reprovados.append([Nome, Curso, Idade, Nota, Matrícula])
        repro+=1

# Depois do loop, formatar todas as datas de uma vez
for celula in aprovados['E'][1:]:  # Pula o cabeçalho
    celula.number_format = "DD/MM/YY"

for celula in reprovados['E'][1:]:
    celula.number_format = "DD/MM/YY"

print(f"Aprovados: {apro}")
print(f"Reprovados: {repro}")
print(f"Média da turma {total/cont:.2f}")
print(f"Aluno com a maior nota: {nome}")

os.makedirs('Projeto Academico/planilhas_criadas', exist_ok=True)

re.save('Projeto Academico/planilhas_criadas/reprovados.xlsx')
ap.save('Projeto Academico/planilhas_criadas/aprovados.xlsx')